from pathlib import Path
from tkinter import Tk, Canvas, Entry, Button, PhotoImage, messagebox
import openpyxl
import re
import sys

# Determine if the script is running in a PyInstaller bundle
if getattr(sys, "frozen", False):
    OUTPUT_PATH = Path.cwd()
    ASSETS_PATH = Path(sys._MEIPASS) / "assets/frame0"
else:
    OUTPUT_PATH = Path(__file__).parent
    ASSETS_PATH = OUTPUT_PATH / "assets/frame0"


def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)


def validate_email(email):
    # Simple regex for email validation
    return re.match(r"[^@]+@[^@]+\.[^@]+", email) is not None


def save_to_excel(data):
    file_path = OUTPUT_PATH / "submissions.xlsx"

    if not file_path.exists():
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "ID",
                "First Name",
                "Last Name",
                "Phone No.",
                "Company",
                "E-Mail",
                "Job Title",
            ]
        )
        workbook.save(file_path)
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

    current_id = sheet.max_row  # ID will be the current row number
    data.insert(0, current_id)  # Insert ID at the beginning of the data list
    sheet.append(data)
    workbook.save(file_path)


def submit_form():
    first_name = entry_1.get()
    last_name = entry_3.get()
    phone_no = entry_2.get()
    company = entry_4.get()
    email = entry_5.get()
    job_title = entry_6.get()

    if not (first_name and last_name and phone_no and company and email and job_title):
        messagebox.showerror("Error", "All fields are required.")
        return

    if not validate_email(email):
        messagebox.showerror("Error", "Invalid email format.")
        return

    data = [first_name, last_name, phone_no, company, email, job_title]
    save_to_excel(data)
    messagebox.showinfo("Success", "Submission saved successfully!")

    # Clear all fields after successful submission
    entry_1.delete(0, "end")
    entry_2.delete(0, "end")
    entry_3.delete(0, "end")
    entry_4.delete(0, "end")
    entry_5.delete(0, "end")
    entry_6.delete(0, "end")


window = Tk()

window.geometry("650x350")
window.configure(bg="#FFFFFF")

canvas = Canvas(
    window,
    bg="#FFFFFF",
    height=350,
    width=650,
    bd=0,
    highlightthickness=0,
    relief="ridge",
)

canvas.place(x=0, y=0)
canvas.create_rectangle(0.0, 88.0, 650.0, 350.0, fill="#243B55", outline="")

canvas.create_rectangle(0.0, 0.0, 650.0, 88.0, fill="#D9D9D9", outline="")

image_image_1 = PhotoImage(file=relative_to_assets("image_1.png"))
image_1 = canvas.create_image(67.0, 44.0, image=image_image_1)

canvas.create_text(
    138.0,
    25.0,
    anchor="nw",
    text="Conference Registration Form",
    fill="#000000",
    font=("Inter Bold", 32 * -1),
)

canvas.create_text(
    23.0,
    113.0,
    anchor="nw",
    text="First Name:",
    fill="#FFFFFF",
    font=("Inter Medium", 20 * -1),
)

canvas.create_text(
    343.0,
    113.0,
    anchor="nw",
    text="Phone No.:",
    fill="#FFFFFF",
    font=("Inter Medium", 20 * -1),
)

canvas.create_text(
    25.0,
    163.0,
    anchor="nw",
    text="Last Name:",
    fill="#FFFFFF",
    font=("Inter Medium", 20 * -1),
)

canvas.create_text(
    351.0,
    163.0,
    anchor="nw",
    text="Company:",
    fill="#FFFFFF",
    font=("Inter Medium", 20 * -1),
)

canvas.create_text(
    66.0,
    213.0,
    anchor="nw",
    text="E-Mail:",
    fill="#FFFFFF",
    font=("Inter Medium", 20 * -1),
)

canvas.create_text(
    359.0,
    213.0,
    anchor="nw",
    text="Job Title:",
    fill="#FFFFFF",
    font=("Inter Medium", 20 * -1),
)

entry_image_1 = PhotoImage(file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(228.5, 125.5, image=entry_image_1)
entry_1 = Entry(bd=0, bg="#FFF9C4", fg="#000716", highlightthickness=0)
entry_1.place(x=149.0, y=110.0, width=159.0, height=29.0)

entry_image_2 = PhotoImage(file=relative_to_assets("entry_2.png"))
entry_bg_2 = canvas.create_image(548.5, 125.5, image=entry_image_2)
entry_2 = Entry(bd=0, bg="#FFF9C4", fg="#000716", highlightthickness=0)
entry_2.place(x=469.0, y=110.0, width=159.0, height=29.0)

entry_image_3 = PhotoImage(file=relative_to_assets("entry_3.png"))
entry_bg_3 = canvas.create_image(228.5, 175.5, image=entry_image_3)
entry_3 = Entry(bd=0, bg="#FFF9C4", fg="#000716", highlightthickness=0)
entry_3.place(x=149.0, y=160.0, width=159.0, height=29.0)

entry_image_4 = PhotoImage(file=relative_to_assets("entry_4.png"))
entry_bg_4 = canvas.create_image(548.5, 175.5, image=entry_image_4)
entry_4 = Entry(bd=0, bg="#FFF9C4", fg="#000716", highlightthickness=0)
entry_4.place(x=469.0, y=160.0, width=159.0, height=29.0)

entry_image_5 = PhotoImage(file=relative_to_assets("entry_5.png"))
entry_bg_5 = canvas.create_image(228.5, 225.5, image=entry_image_5)
entry_5 = Entry(bd=0, bg="#FFF9C4", fg="#000716", highlightthickness=0)
entry_5.place(x=149.0, y=210.0, width=159.0, height=29.0)

entry_image_6 = PhotoImage(file=relative_to_assets("entry_6.png"))
entry_bg_6 = canvas.create_image(548.5, 225.5, image=entry_image_6)
entry_6 = Entry(bd=0, bg="#FFF9C4", fg="#000716", highlightthickness=0)
entry_6.place(x=469.0, y=210.0, width=159.0, height=29.0)

button_image_1 = PhotoImage(file=relative_to_assets("button_1.png"))
button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=submit_form,
    relief="flat",
)
button_1.place(x=221.0, y=272.0, width=207.0, height=56.0)

window.resizable(False, False)
window.mainloop()
